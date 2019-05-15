##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2019 Université catholique de Louvain
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font
from internship import models
from internship.models.internship_score import InternshipScore
from internship.models.internship_score_mapping import InternshipScoreMapping
from internship.utils.exporting.spreadsheet import columns_resizing, add_row


def export_xls(cohort):
    workbook = Workbook()
    worksheet = workbook.active
    _add_header(cohort, worksheet)
    columns_resizing(worksheet, {'A': 32, 'B': 16, 'C': 11, 'D': 5, 'E': 7, 'F': 5, 'G': 7, 'H': 5, 'I': 7, 'J': 5,
                                 'K': 7, 'L': 5, 'M': 7, 'N': 5, 'O': 7, 'P': 5, 'Q': 7, 'R': 5, 'S': 7, 'T': 5,
                                 'U': 7, 'V': 5, 'W': 7, 'X': 5, 'Y': 7, 'Z': 5, 'AA': 7})
    _add_students(cohort, worksheet)
    return save_virtual_workbook(workbook)


def export_xls_with_scores(cohort, periods, students, internships):
    workbook = Workbook()
    worksheet = workbook.active
    _add_header(cohort, worksheet)
    _make_complete_list(periods, students, worksheet)
    _make_internship_sheets(internships, periods, students, workbook)
    return save_virtual_workbook(workbook)


def _make_internship_sheets(internships, periods, students, workbook):
    for internship in sorted(internships):
        workbook.create_sheet(internship)
        worksheet = workbook[internship]
        _add_sheet_header(worksheet)
        _add_sheet_content(internship, periods, students, worksheet)


def _add_sheet_content(internship, periods, students, worksheet):
    for student in students:
        columns = []
        columns.append(student.person.last_name.upper())
        columns.append(student.person.first_name)
        columns.append(student.registration_id)
        _complete_student_row_by_internship(columns, internship, periods, student)
        add_row(worksheet, columns)


def _complete_student_row_by_internship(columns, internship, periods, student):
    for period in periods:
        if period.name in student.specialties.keys() and student.specialties[period.name] == internship:
            _append_row_data(columns, period, student)


def _append_row_data(columns, period, student):
    if period.name in student.organizations.keys():
        columns.append(student.organizations[period.name])
    if period.name in student.periods_scores.keys():
        columns.append(student.periods_scores[period.name])
    else:
        columns.append('')


def _make_complete_list(periods, students, worksheet):
    columns_resizing(worksheet, {'A': 32, 'B': 16, 'C': 11, 'D': 7, 'E': 7, 'F': 7, 'G': 7, 'H': 7, 'I': 7, 'J': 7,
                                 'K': 7, 'L': 7, 'M': 7, 'N': 7, 'O': 7, 'P': 7, 'Q': 7, 'R': 7, 'S': 7, 'T': 7,
                                 'U': 7, 'V': 7, 'W': 7, 'X': 7, 'Y': 7, 'Z': 7, 'AA': 7, 'AB': 7, 'AC': 7, 'AD': 7,
                                 'AE': 7, 'AF': 7, 'AG': 7, 'AH': 7, 'AI': 7, 'AJ': 7, 'AK': 7, 'AL': 7, 'AM': 7,
                                 'AN': 7, 'AO': 7, 'AP': 7, 'AQ': 7})
    for student in students:
        columns = []
        columns.append(student.person.last_name.upper())
        columns.append(student.person.first_name)
        columns.append(student.registration_id)
        _complete_student_row_for_all_internships(columns, periods, student)
        add_row(worksheet, columns)


def _complete_student_row_for_all_internships(columns, periods, student):
    for period in periods:
        if period.name in student.specialties.keys():
            columns.append(student.specialties[period.name])
        _append_row_data(columns, period, student)


def _add_sheet_header(worksheet):
    column_titles = ["Nom", "Prénom", "NOMA", "LIEU", "COTE"]
    add_row(worksheet, column_titles)
    cells = worksheet.range("A1:AAA1")[0]
    for cell in cells:
        cell.font = Font(bold=True)


def _add_header(cohort, worksheet):
    periods = models.period.find_by_cohort(cohort)
    column_titles = ["Nom", "Prénom", "NOMA"]
    for period in periods:
        column_titles.append(period.name)
        column_titles.append("{}+".format(period.name))
        column_titles.append("{}-Score".format(period.name))
    add_row(worksheet, column_titles)
    cells = worksheet.range("A1:AAA1")[0]
    for cell in cells:
        cell.font = Font(bold=True)


def _add_students(cohort, worksheet):
    students_stat = models.internship_student_affectation_stat.find_by_cohort(cohort)
    previous_student = None
    columns = []
    for student_stat in students_stat:
        # Line breaking condition.
        if student_stat.student.registration_id != previous_student:
            if len(columns) > 0:
                add_row(worksheet, columns)
                columns = []
            columns.append(student_stat.student.person.last_name.upper())
            columns.append(student_stat.student.person.first_name)
            columns.append(student_stat.student.registration_id)
            previous_student = student_stat.student.registration_id
        _add_period(columns, student_stat)
    add_row(worksheet, columns)


def _add_period(columns, student_stat):
    _deal_with_empty_periods(columns, student_stat)
    columns.append(student_stat.speciality.acronym_with_sequence())
    columns.append("{}{}".format(student_stat.speciality.acronym, student_stat.organization.reference))


def _deal_with_empty_periods(columns, student_stat):
    period_position = (student_stat.period.number() * 2) + 1  # Consider the first columns and 2 columns per period.
    if len(columns) < period_position:
        diff = period_position - len(columns)
        while diff > 0:  # Add empty columns when there is no data to the periods before the current one.
            columns.append("")
            columns.append("")
            diff -= 2  # Subtracts the number of columns appended within the while loop.
